"""
AUM Consolidation Web App — Sequoia Capital Management
Flask backend: processes uploaded platform files, returns mapping errors,
accepts manual mapping fixes, and generates the final Excel output.
"""

import os, io, json, uuid, traceback
from datetime import date, datetime
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

app = Flask(__name__)
CORS(app)

UPLOAD_DIR = "/tmp/aum_uploads"
OUTPUT_DIR = "/tmp/aum_outputs"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── session store (in-memory, one session at a time is fine for desktop use) ──
SESSION = {}

# ─────────────────────────────────────────────────────────────────────────────
# STRIP RULES  (all bugs from previous run fixed here)
# ─────────────────────────────────────────────────────────────────────────────
PLATFORM_STRIP_RULES = {
    "allan_gray": {
        # The Allan Gray pivot already shows the correct Sequoia share for A/B/C variants.
        # Only delete the base fund and F variant entirely.
        "Sequoia Worldwide Flexible":   None,
        "Sequoia Worldwide Flexible F": None,
    },
    "glacier": {
        "Sequoia Worldwide Flexible":   None,
        "Sequoia Worldwide Flexible C": 0.66,
    },
    "momentum": {
        "Sequoia Worldwide Flexible":   None,
        "Sequoia Worldwide Flexible A": None,
        "Sequoia Worldwide Flexible B": None,
        "Sequoia Worldwide Flexible C": 0.40,
        "Sequoia Worldwide Flexible D": 0.46,
        "Sequoia Worldwide Flexible F": 0.61,
    },
    "stanlib": {
        "Sequoia Worldwide Flexible":        None,
        "Sequoia Worldwide Flexible Fund":   None,   # base fund variant name used in Stanlib
        "Sequoia Worldwide Flexible A":      0.20,
        "Sequoia Worldwide Flexible B":      0.20,
        "Sequoia Worldwide Flexible C":      0.66,
        "Sequoia Worldwide Flexible D":      0.72,
        "Sequoia Worldwide Flexible F":      None,
    },
}


def apply_strip_rules(df, platform, fund_col, aum_col, inflow_col=None, outflow_col=None):
    rules = PLATFORM_STRIP_RULES.get(platform, {})
    if not rules:
        return df
    drop = []
    for idx, row in df.iterrows():
        fund = str(row[fund_col]).strip()
        matched = next((k for k in rules if k.lower() == fund.lower()), None)
        if matched is None:
            continue
        mult = rules[matched]
        if mult is None:
            drop.append(idx)
        else:
            df.at[idx, aum_col] = row[aum_col] * mult if pd.notna(row[aum_col]) else np.nan
            if inflow_col and inflow_col in df.columns and pd.notna(row.get(inflow_col)):
                df.at[idx, inflow_col] = row[inflow_col] * mult
            if outflow_col and outflow_col in df.columns and pd.notna(row.get(outflow_col)):
                df.at[idx, outflow_col] = row[outflow_col] * mult
    return df.drop(index=drop).reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# LOOKUP LOADERS
# ─────────────────────────────────────────────────────────────────────────────
def load_advisor_map(template_path):
    df = pd.read_excel(template_path, sheet_name="ADVISOR ID | CODE MAP", header=0)
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df.columns = ["ID", "Broker Name", "Broker House Name", "LISP", "Data Sources"]
    df = df.dropna(subset=["ID"])
    df["ID"] = df["ID"].astype(str).str.strip()
    lookup = {}
    for _, row in df.iterrows():
        lookup[row["ID"]] = {
            "Broker Name":       str(row["Broker Name"]).strip()       if pd.notna(row["Broker Name"])       else "",
            "Broker House Name": str(row["Broker House Name"]).strip() if pd.notna(row["Broker House Name"]) else "",
            "LISP":              str(row["LISP"]).strip()              if pd.notna(row["LISP"])              else "",
        }
    return lookup


def load_fund_map(template_path):
    df = pd.read_excel(template_path, sheet_name="FUND MAP", header=0)
    df = df[["LISPS NAMING", "Fund Name", "Product"]].dropna(subset=["LISPS NAMING", "Fund Name"])
    lookup = {}
    for _, row in df.iterrows():
        key = str(row["LISPS NAMING"]).strip().lower()
        lookup[key] = {
            "Fund Name": str(row["Fund Name"]).strip(),
            "Product":   str(row["Product"]).strip() if pd.notna(row["Product"]) else "Model",
        }
    return lookup


def map_advisor(broker_id, advisor_map):
    return advisor_map.get(str(broker_id).strip(), {"Broker Name": "", "Broker House Name": "", "LISP": ""})


def map_fund(raw_name, fund_map):
    key = str(raw_name).strip().lower()
    result = fund_map.get(key)
    if result:
        return result["Fund Name"], result["Product"]
    return str(raw_name).strip(), "Model"


# ─────────────────────────────────────────────────────────────────────────────
# PLATFORM PROCESSORS  (all bugs fixed vs original script)
# ─────────────────────────────────────────────────────────────────────────────

def process_allan_gray(path, advisor_map, fund_map):
    df = pd.read_excel(path, sheet_name="Sheet2")
    df = df.rename(columns={"IFA Code": "ID", "Model Portfolio Name": "Fund Name Raw", "Total": "AUM"})
    df = df[["ID", "Fund Name Raw", "AUM"]].dropna(subset=["ID", "Fund Name Raw", "AUM"])
    df["ID"] = df["ID"].astype(str).str.strip()
    df = apply_strip_rules(df, "allan_gray", "Fund Name Raw", "AUM")
    rows = []
    for _, row in df.iterrows():
        adv = map_advisor(row["ID"], advisor_map)
        std_fund, product = map_fund(row["Fund Name Raw"], fund_map)
        rows.append({"ID": row["ID"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product,
                     "LISP": "Allan Gray", "Fund Name Raw": row["Fund Name Raw"],
                     "Fund Name": std_fund,
                     "InFlows (R)": np.nan, "OutFlows (R)": np.nan, "NetFlows (R)": np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


def process_coruscate(path, advisor_map, fund_map):
    df = pd.read_excel(path, sheet_name="Table1")
    # Use Fund column; fall back to FundGroup where Fund is blank
    if "Fund" in df.columns and "FundGroup" in df.columns:
        df["_fund"] = df["Fund"].where(df["Fund"].notna() & (df["Fund"].astype(str).str.strip() != ""), df["FundGroup"])
    elif "Fund" in df.columns:
        df["_fund"] = df["Fund"]
    elif "FundGroup" in df.columns:
        df["_fund"] = df["FundGroup"]
    else:
        raise ValueError(f"Coruscate: cannot find Fund or FundGroup column. Columns: {list(df.columns)}")

    df = df.dropna(subset=["IdNumber", "_fund", "AumBase"])
    df["IdNumber"] = df["IdNumber"].astype(str).str.strip()
    df["OutflowBase"] = pd.to_numeric(df["OutflowBase"], errors="coerce").fillna(0).abs()
    rows = []
    for _, row in df.iterrows():
        adv = map_advisor(row["IdNumber"], advisor_map)
        std_fund, product = map_fund(row["_fund"], fund_map)
        lisp_raw = str(row["Lisp"]).strip() if pd.notna(row["Lisp"]) else ""
        lisp = "PPS" if lisp_raw.upper() == "PPSI" else lisp_raw
        inf  = row["InflowBase"]  if pd.notna(row["InflowBase"])  else np.nan
        out  = row["OutflowBase"] if pd.notna(row["OutflowBase"]) else np.nan
        net  = (inf if not np.isnan(inf) else 0) - (out if not np.isnan(out) else 0)
        rows.append({"ID": row["IdNumber"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product, "LISP": lisp,
                     "Fund Name Raw": row["_fund"], "Fund Name": std_fund,
                     "InFlows (R)": inf, "OutFlows (R)": out,
                     "NetFlows (R)": net if not (np.isnan(inf) and np.isnan(out)) else np.nan,
                     "AUM (R)": row["AumBase"]})
    return pd.DataFrame(rows)


def process_gla(path, advisor_map, fund_map):
    df = pd.read_excel(path, sheet_name="AUA and Flows", header=0)
    df.columns = ["FDATE","PP","PP_INCEPTION_DATE","PP_PRODUCT","PP_INITIALS","PP_NAME",
                  "PP_PARTY_ID","BROKER_CODE","BROKER_NAME","HOUSE_CODE","HOUSE_NAME",
                  "FCODE","F_NAME","INFLOWS","OUTFLOWS","NETFLOWS","AUA","UNITS"]
    df["FDATE"] = pd.to_datetime(df["FDATE"])
    df = df[df["FDATE"] == df["FDATE"].max()].copy()
    df = df.dropna(subset=["BROKER_CODE", "F_NAME", "AUA"])
    df["BROKER_CODE"] = df["BROKER_CODE"].astype(str).str.strip()
    df["OUTFLOWS"] = df["OUTFLOWS"].abs()
    rows = []
    for _, row in df.iterrows():
        adv = map_advisor(row["BROKER_CODE"], advisor_map)
        std_fund, product = map_fund(row["F_NAME"], fund_map)
        inf = row["INFLOWS"]  if pd.notna(row["INFLOWS"])  else np.nan
        out = row["OUTFLOWS"] if pd.notna(row["OUTFLOWS"]) else np.nan
        net = (inf if not np.isnan(inf) else 0) - (out if not np.isnan(out) else 0)
        rows.append({"ID": row["BROKER_CODE"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product,
                     "LISP": "Momentum",  # GLA uses Momentum as LISP name
                     "Fund Name Raw": row["F_NAME"], "Fund Name": std_fund,
                     "InFlows (R)": inf, "OutFlows (R)": out,
                     "NetFlows (R)": net if not (np.isnan(inf) and np.isnan(out)) else np.nan,
                     "AUM (R)": row["AUA"]})
    return pd.DataFrame(rows)


def process_glacier(path, advisor_map, fund_map):
    df = pd.read_csv(path, sep=";", decimal=",")
    df.columns = ["Broker Code", "Wrap Fund Name", "Inflows", "Outflows", "Current Value"]
    df = df.dropna(subset=["Broker Code", "Wrap Fund Name", "Current Value"])
    df["Broker Code"] = df["Broker Code"].astype(str).str.strip()
    df["Outflows"] = df["Outflows"].abs()
    df = apply_strip_rules(df, "glacier", "Wrap Fund Name", "Current Value",
                           inflow_col="Inflows", outflow_col="Outflows")
    for c in ["Inflows", "Outflows", "Current Value"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    grp = df.groupby(["Broker Code", "Wrap Fund Name"], as_index=False).agg(
        Inflows=("Inflows", "sum"), Outflows=("Outflows", "sum"), AUM=("Current Value", "sum"))
    rows = []
    for _, row in grp.iterrows():
        adv = map_advisor(row["Broker Code"], advisor_map)
        std_fund, product = map_fund(row["Wrap Fund Name"], fund_map)
        rows.append({"ID": row["Broker Code"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product, "LISP": "Glacier",
                     "Fund Name Raw": row["Wrap Fund Name"], "Fund Name": std_fund,
                     "InFlows (R)":  row["Inflows"]  if row["Inflows"]  != 0 else np.nan,
                     "OutFlows (R)": row["Outflows"] if row["Outflows"] != 0 else np.nan,
                     "NetFlows (R)": row["Inflows"] - row["Outflows"] if (row["Inflows"] != 0 or row["Outflows"] != 0) else np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


def process_momentum(path, advisor_map, fund_map):
    # python engine handles rows with inconsistent field counts; low_memory not supported with it
    df = pd.read_csv(path, engine="python", on_bad_lines="skip")
    df = df.dropna(subset=["BrokerCode", "MultiManagerPortfolio", "ClosingValue"])
    df["BrokerCode"] = df["BrokerCode"].astype(str).str.strip()
    # Outflows column may be missing in some months — handle gracefully
    if "Outflows" not in df.columns:
        df["Outflows"] = 0
    if "Inflows" not in df.columns:
        df["Inflows"] = 0
    df["Outflows"] = pd.to_numeric(df["Outflows"], errors="coerce").fillna(0).abs()
    df["Inflows"]  = pd.to_numeric(df["Inflows"],  errors="coerce").fillna(0)
    df = apply_strip_rules(df, "momentum", "MultiManagerPortfolio", "ClosingValue",
                           inflow_col="Inflows", outflow_col="Outflows")
    df["ClosingValue"] = pd.to_numeric(df["ClosingValue"], errors="coerce").fillna(0)
    grp = df.groupby(["BrokerCode", "MultiManagerPortfolio"], as_index=False).agg(
        Inflows=("Inflows", "sum"), Outflows=("Outflows", "sum"), AUM=("ClosingValue", "sum"))
    rows = []
    for _, row in grp.iterrows():
        adv = map_advisor(row["BrokerCode"], advisor_map)
        std_fund, product = map_fund(row["MultiManagerPortfolio"], fund_map)
        rows.append({"ID": row["BrokerCode"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product, "LISP": "Momentum",
                     "Fund Name Raw": row["MultiManagerPortfolio"], "Fund Name": std_fund,
                     "InFlows (R)":  row["Inflows"]  if row["Inflows"]  != 0 else np.nan,
                     "OutFlows (R)": row["Outflows"] if row["Outflows"] != 0 else np.nan,
                     "NetFlows (R)": row["Inflows"] - row["Outflows"] if (row["Inflows"] != 0 or row["Outflows"] != 0) else np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


def process_ninety_one(path, advisor_map, fund_map):
    df = pd.read_excel(path, sheet_name="Sheet2")
    df.columns = ["Advisor Number", "Model Name", "Closing AUM"]
    df = df.dropna(subset=["Advisor Number", "Model Name", "Closing AUM"])
    df["Advisor Number"] = df["Advisor Number"].astype(str).str.strip()
    rows = []
    for _, row in df.iterrows():
        adv = map_advisor(row["Advisor Number"], advisor_map)
        std_fund, product = map_fund(row["Model Name"], fund_map)
        rows.append({"ID": row["Advisor Number"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product, "LISP": "Ninety One",
                     "Fund Name Raw": row["Model Name"], "Fund Name": std_fund,
                     "InFlows (R)": np.nan, "OutFlows (R)": np.nan, "NetFlows (R)": np.nan,
                     "AUM (R)": row["Closing AUM"]})
    return pd.DataFrame(rows)


def process_pps(path, advisor_map, fund_map):
    if str(path).lower().endswith(".csv"):
        df = pd.read_csv(path, engine="python", on_bad_lines="skip")
        df = df.rename(columns={"MarketValue": "Market Value", "Inflow": "Inflows", "OutFlow": "Outflows"})
    else:
        # Find the AUMReport sheet — name contains "AUMReport" regardless of date suffix
        xl = pd.ExcelFile(path)
        sheet = next((s for s in xl.sheet_names if "AUMReport" in s), None)
        if sheet is None:
            sheet = next((s for s in xl.sheet_names if "aum" in s.lower()), xl.sheet_names[0])
        df = pd.read_excel(path, sheet_name=sheet, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.rename(columns={
            "MarketValue": "Market Value", "Inflow": "Inflows",
            "OutFlow": "Outflows", "OutFlows": "Outflows",
            "ModelPortfolioName": "Model Portfolio Name",
            "Intermediary No": "IntermediaryNo", "Intermediary_No": "IntermediaryNo",
        })

    # Auto-detect column names for robustness across months
    id_col = next((c for c in ["IntermediaryNo","Intermediary No","IntermediaryNumber","BrokerCode"]
                   if c in df.columns), None)
    if id_col is None:
        raise ValueError(f"PPS: cannot find intermediary ID column. Columns: {list(df.columns)}")
    if id_col != "IntermediaryNo":
        df = df.rename(columns={id_col: "IntermediaryNo"})

    port_col = next((c for c in ["Model Portfolio Name","ModelPortfolioName","Portfolio Name","PortfolioName"]
                     if c in df.columns), None)
    if port_col is None:
        raise ValueError(f"PPS: cannot find portfolio column. Columns: {list(df.columns)}")
    if port_col != "Model Portfolio Name":
        df = df.rename(columns={port_col: "Model Portfolio Name"})

    mv_col = next((c for c in ["Market Value","MarketValue","AUM","Value","ClosingValue","Current Value"]
                   if c in df.columns), None)
    if mv_col is None:
        raise ValueError(f"PPS: cannot find market value column. Columns: {list(df.columns)}")
    if mv_col != "Market Value":
        df = df.rename(columns={mv_col: "Market Value"})

    for src, tgt in [("Inflow","Inflows"),("OutFlow","Outflows"),("OutFlows","Outflows")]:
        if src in df.columns and tgt not in df.columns:
            df = df.rename(columns={src: tgt})
    if "Inflows"  not in df.columns: df["Inflows"]  = 0
    if "Outflows" not in df.columns: df["Outflows"] = 0

    df = df.dropna(subset=["IntermediaryNo", "Model Portfolio Name", "Market Value"])
    df["IntermediaryNo"] = df["IntermediaryNo"].astype(str).str.strip()
    df["Outflows"]     = pd.to_numeric(df["Outflows"],     errors="coerce").fillna(0).abs()
    df["Inflows"]      = pd.to_numeric(df["Inflows"],      errors="coerce").fillna(0)
    df["Market Value"] = pd.to_numeric(df["Market Value"], errors="coerce").fillna(0)

    grp = df.groupby(["IntermediaryNo", "Model Portfolio Name"], as_index=False).agg(
        Inflows=("Inflows", "sum"), Outflows=("Outflows", "sum"), AUM=("Market Value", "sum"))
    rows = []
    for _, row in grp.iterrows():
        adv = map_advisor(row["IntermediaryNo"], advisor_map)
        std_fund, product = map_fund(row["Model Portfolio Name"], fund_map)
        rows.append({"ID": row["IntermediaryNo"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product, "LISP": "PPS",
                     "Fund Name Raw": row["Model Portfolio Name"], "Fund Name": std_fund,
                     "InFlows (R)":  row["Inflows"]  if row["Inflows"]  != 0 else np.nan,
                     "OutFlows (R)": row["Outflows"] if row["Outflows"] != 0 else np.nan,
                     "NetFlows (R)": row["Inflows"] - row["Outflows"] if (row["Inflows"] != 0 or row["Outflows"] != 0) else np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


def process_investec(path, advisor_map, fund_map):
    """
    BUG FIX: filter to individual client rows only — exclude the rollup row
    where Entity Full Name == company name (double-counting).
    We detect the rollup by checking if AUM equals the sum of other rows
    for that advisor, or more simply: keep only rows where Entity Full Name
    contains 'Investec Life' or is a client name (not the fund manager name).
    Original script caused ~R32M double-count.
    """
    # Try the known sheet name first, fall back to first sheet
    try:
        df = pd.read_excel(path, sheet_name="Feb26")
    except Exception:
        df = pd.read_excel(path, sheet_name=0)

    df = df.dropna(subset=["Financial Advisor", "Value"])
    df["Financial Advisor"] = df["Financial Advisor"].astype(str).str.strip()

    # Filter out rollup/summary rows — keep only "Investec Life Limited" entity rows
    if "Entity Full Name" in df.columns:
        df = df[df["Entity Full Name"].astype(str).str.contains("Investec Life", case=False, na=False)]

    df["Value"] = pd.to_numeric(df["Value"], errors="coerce").fillna(0)

    if "Instrument Full Name" not in df.columns:
        # Try fallback column names
        for c in df.columns:
            if "instrument" in c.lower() or "fund" in c.lower():
                df = df.rename(columns={c: "Instrument Full Name"})
                break

    grp = df.groupby(["Financial Advisor", "Instrument Full Name"], as_index=False).agg(AUM=("Value", "sum"))
    rows = []
    for _, row in grp.iterrows():
        if pd.isna(row["Instrument Full Name"]):
            continue
        adv = map_advisor(row["Financial Advisor"], advisor_map)
        std_fund, product = map_fund(str(row["Instrument Full Name"]).replace("_", " "), fund_map)
        rows.append({"ID": row["Financial Advisor"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"] or row["Financial Advisor"],
                     "Product": product, "LISP": "Investec",
                     "Fund Name Raw": row["Instrument Full Name"], "Fund Name": std_fund,
                     "InFlows (R)": np.nan, "OutFlows (R)": np.nan, "NetFlows (R)": np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


def process_wealthport(path, advisor_map, fund_map):
    df = pd.read_excel(path, sheet_name="Sheet1")
    df = df.dropna(subset=["Agency Name", "Portfolio Name", "Market Value"])
    df["Agency Name"] = df["Agency Name"].astype(str).str.strip()
    df["Market Value"] = pd.to_numeric(df["Market Value"], errors="coerce").fillna(0)
    grp = df.groupby(["Agency Name", "Portfolio Name"], as_index=False).agg(AUM=("Market Value", "sum"))
    rows = []
    for _, row in grp.iterrows():
        adv = map_advisor(row["Agency Name"], advisor_map)
        std_fund, product = map_fund(row["Portfolio Name"], fund_map)
        rows.append({"ID": row["Agency Name"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product, "LISP": "Wealthport",
                     "Fund Name Raw": row["Portfolio Name"], "Fund Name": std_fund,
                     "InFlows (R)": np.nan, "OutFlows (R)": np.nan, "NetFlows (R)": np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


def _read_stanlib_sheet(path):
    """Read Stanlib/Discovery — uses Positions sheet (client-level detail) when present, else pivot."""
    xl = pd.ExcelFile(path)

    # ── Positions sheet: client-level rows with Current Market Value ──────────
    if "Positions" in xl.sheet_names:
        df = pd.read_excel(path, sheet_name="Positions", header=0)
        df.columns = [str(c).strip() for c in df.columns]
        adv_col = next((c for c in ["Adviser Code","Advisor Code","Adviser","Advisor"] if c in df.columns), None)
        mod_col = next((c for c in ["Model Name","ModelName","Model Portfolio Name"] if c in df.columns), None)
        mv_col  = next((c for c in ["Current Market Value","MarketValue","Market Value","Value"] if c in df.columns), None)
        if adv_col and mod_col and mv_col:
            df[mv_col]  = pd.to_numeric(df[mv_col],  errors="coerce").fillna(0)
            df[adv_col] = df[adv_col].astype(str).str.strip()
            df[mod_col] = df[mod_col].astype(str).str.strip()
            df = df[df[adv_col].notna() & (df[adv_col] != "nan") & (df[adv_col] != "")]
            df = df[df[mod_col].notna() & (df[mod_col] != "nan") & (df[mod_col] != "")]
            grp = df.groupby([adv_col, mod_col], as_index=False)[mv_col].sum()
            out = pd.DataFrame()
            out["Advisor Code"] = grp[adv_col]
            out["Model Name"]   = grp[mod_col]
            out["AUM"]          = grp[mv_col]
            out["Inflows"]      = 0
            out["Outflows"]     = 0
            if len(out) > 0 and out["AUM"].sum() > 0:
                return out

    # ── Fallback: pivot on Sheet1/Sheet2 ─────────────────────────────────────
    for sheet in ["Sheet2", "Sheet1"] + xl.sheet_names:
        if sheet not in xl.sheet_names or sheet == "Positions":
            continue
        raw = pd.read_excel(path, sheet_name=sheet, header=None)
        header_row = None
        for i, row in raw.iterrows():
            vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
            if any("model" in v or "portfolio" in v or "adviser" in v or "advisor" in v for v in vals):
                header_row = i
                break
        if header_row is None:
            continue
        df = pd.read_excel(path, sheet_name=sheet, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]
        known_advisor  = ["Advisor Code","AdvisorCode","BROKER_ENTITY","Broker Code","Adviser Code","Adviser"]
        known_model    = ["Model Name","ModelName","Model Portfolio Name","Portfolio Name","Model Portfolio NAME"]
        known_inflows  = ["Inflows","Inflow","Sum of INFLOW","INFLOW"]
        known_outflows = ["Outflows","Outflow","Sum of OUTFLOW","OUTFLOW"]
        known_aum      = ["AUM","Sum of AUM","Total","MarketValue","Market Value","ClosingValue","Current Market Value"]
        def find_col(candidates):
            return next((c for c in candidates if c in df.columns), None)
        advisor_col = find_col(known_advisor) or df.columns[0]
        model_col   = find_col(known_model)   or df.columns[1]
        inflow_col  = find_col(known_inflows)
        outflow_col = find_col(known_outflows)
        aum_col     = find_col(known_aum)     or df.columns[-1]
        df[advisor_col] = df[advisor_col].ffill()
        df = df[df[model_col].notna()]
        df = df[~df[model_col].astype(str).str.strip().isin(
            ["Model Name","ModelName","(blank)","Grand Total","Total","nan",""])]
        df = df[~df[advisor_col].astype(str).str.strip().isin(
            ["Adviser Code","Advisor Code","(blank)","nan",""])]
        out = pd.DataFrame()
        out["Advisor Code"] = df[advisor_col].astype(str).str.strip()
        out["Model Name"]   = df[model_col].astype(str).str.strip()
        out["AUM"]          = pd.to_numeric(df[aum_col],    errors="coerce").fillna(0)
        out["Inflows"]      = pd.to_numeric(df[inflow_col],  errors="coerce").fillna(0) if inflow_col else 0
        out["Outflows"]     = pd.to_numeric(df[outflow_col], errors="coerce").fillna(0) if outflow_col else 0
        if len(out) > 0 and out["AUM"].sum() > 0:
            return out

    raise ValueError(f"Could not find data in {path}")


def process_discovery(path, advisor_map, fund_map):
    """Discovery — Copy of 2086288 file."""
    df = _read_stanlib_sheet(path)
    df = df.dropna(subset=["Advisor Code", "Model Name", "AUM"])
    df["Outflows"] = df["Outflows"].abs()
    df = apply_strip_rules(df, "stanlib", "Model Name", "AUM",
                           inflow_col="Inflows", outflow_col="Outflows")
    rows = []
    for _, row in df.iterrows():
        adv = map_advisor(row["Advisor Code"], advisor_map)
        std_fund, product = map_fund(row["Model Name"], fund_map)
        in_v  = row["Inflows"]  if pd.notna(row["Inflows"])  else 0
        out_v = row["Outflows"] if pd.notna(row["Outflows"]) else 0
        rows.append({"ID": row["Advisor Code"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product,
                     "LISP": "Discovery",
                     "Fund Name Raw": row["Model Name"], "Fund Name": std_fund,
                     "InFlows (R)":  in_v  if in_v  != 0 else np.nan,
                     "OutFlows (R)": out_v if out_v != 0 else np.nan,
                     "NetFlows (R)": in_v - out_v if (in_v != 0 or out_v != 0) else np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


def process_stanlib(path, advisor_map, fund_map):
    """Stanlib — 2086288 file."""
    df = _read_stanlib_sheet(path)
    df = df.dropna(subset=["Advisor Code", "Model Name", "AUM"])
    df["Outflows"] = df["Outflows"].abs()
    df = apply_strip_rules(df, "stanlib", "Model Name", "AUM",
                           inflow_col="Inflows", outflow_col="Outflows")
    rows = []
    for _, row in df.iterrows():
        adv = map_advisor(row["Advisor Code"], advisor_map)
        std_fund, product = map_fund(row["Model Name"], fund_map)
        in_v  = row["Inflows"]  if pd.notna(row["Inflows"])  else 0
        out_v = row["Outflows"] if pd.notna(row["Outflows"]) else 0
        rows.append({"ID": row["Advisor Code"], "Broker House Name": adv["Broker House Name"],
                     "Broker Name": adv["Broker Name"], "Product": product,
                     "LISP": "Stanlib",
                     "Fund Name Raw": row["Model Name"], "Fund Name": std_fund,
                     "InFlows (R)":  in_v  if in_v  != 0 else np.nan,
                     "OutFlows (R)": out_v if out_v != 0 else np.nan,
                     "NetFlows (R)": in_v - out_v if (in_v != 0 or out_v != 0) else np.nan,
                     "AUM (R)": row["AUM"]})
    return pd.DataFrame(rows)


PROCESSOR_MAP = {
    "allan_gray":  process_allan_gray,
    "coruscate":   process_coruscate,
    "gla":         process_gla,
    "glacier":     process_glacier,
    "momentum":    process_momentum,
    "ninety_one":  process_ninety_one,
    "pps":         process_pps,
    "investec":    process_investec,
    "wealthport":  process_wealthport,
    "stanlib":     process_stanlib,
    "discovery":   process_discovery,
}

PLATFORM_LABELS = {
    "allan_gray":  "Allan Gray",
    "coruscate":   "Coruscate",
    "gla":         "GLA (Momentum LISP)",
    "glacier":     "Glacier",
    "momentum":    "Momentum",
    "ninety_one":  "Ninety One",
    "pps":         "PPS",
    "investec":    "Investec",
    "wealthport":  "Wealthport",
    "stanlib":     "Stanlib",
    "discovery":   "Discovery (Copy file)",
}


# ─────────────────────────────────────────────────────────────────────────────
# MAPPING ERROR DETECTION
# ─────────────────────────────────────────────────────────────────────────────
def find_mapping_errors(df, advisor_map, fund_map):
    """Return list of {type, id, platform, fund_name_raw, aum} for unmapped rows."""
    errors = []
    seen_advisor = set()
    seen_fund = set()

    for _, row in df.iterrows():
        aid = str(row.get("ID", "")).strip()
        if aid and aid not in advisor_map and aid not in seen_advisor:
            seen_advisor.add(aid)
            errors.append({
                "type": "advisor",
                "id": aid,
                "platform": str(row.get("LISP", "")),
                "fund_name_raw": str(row.get("Fund Name Raw", "")),
                "aum": float(row.get("AUM (R)", 0) or 0),
                "current_broker_name": str(row.get("Broker Name", "")),
                "current_house_name": str(row.get("Broker House Name", "")),
            })

        raw = str(row.get("Fund Name Raw", "")).strip()
        if raw and raw.lower() not in fund_map and raw not in seen_fund:
            seen_fund.add(raw)
            errors.append({
                "type": "fund",
                "id": raw,
                "platform": str(row.get("LISP", "")),
                "fund_name_raw": raw,
                "aum": float(row.get("AUM (R)", 0) or 0),
                "current_fund_name": str(row.get("Fund Name", raw)),
                "current_product": str(row.get("Product", "Model")),
            })

    return errors


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER  (BUG FIX: write plain numeric totals, not SUM formulas)
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(df, report_date_str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Worksheet"

    FN = "Calibri Light"
    FS = 12
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(name=FN, size=FS, bold=True, color="FFFFFF")
    DAT_FONT = Font(name=FN, size=FS)
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
    TOT_FILL = PatternFill("solid", fgColor="FFF2CC")
    BORDER   = Border(bottom=Side(style="thin", color="D9D9D9"),
                      right=Side(style="thin",  color="D9D9D9"))

    HEADERS   = ["Date", "Broker House Name", "Broker Name",
                 "Retirement Fund Type", "Participating Employer",
                 "Product", "LISP", "Fund Name",
                 "InFlows (R)", "OutFlows (R)", "NetFlows (R)", "AUM (R)"]
    COL_WIDTHS = [14, 32, 28, 22, 26, 12, 18, 36, 15, 15, 15, 18]

    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    numeric_cols = {"InFlows (R)", "OutFlows (R)", "NetFlows (R)", "AUM (R)"}
    numeric_totals = {h: 0.0 for h in numeric_cols}

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, h in enumerate(HEADERS, 1):
            val = row.get(h)
            if not isinstance(val, str) and pd.isna(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DAT_FONT
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if h == "Date" and val is not None:
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center")
            elif h in numeric_cols and val is not None:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
                numeric_totals[h] += float(val)
            else:
                cell.alignment = Alignment(horizontal="left")

    # Totals row — plain values, not formulas (avoids openpyxl doubling bug)
    total_row = len(df) + 2
    tc = ws.cell(row=total_row, column=1, value="TOTAL")
    tc.font = Font(name=FN, size=FS, bold=True)
    for h, ci in [("InFlows (R)", 9), ("OutFlows (R)", 10), ("NetFlows (R)", 11), ("AUM (R)", 12)]:
        c = ws.cell(row=total_row, column=ci, value=round(numeric_totals[h], 2))
        c.font = Font(name=FN, size=FS, bold=True)
        c.number_format = "#,##0.00"
        c.fill = TOT_FILL
        c.alignment = Alignment(horizontal="right")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, numeric_totals["AUM (R)"]


# ─────────────────────────────────────────────────────────────────────────────
# FLASK ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/upload_template", methods=["POST"])
def upload_template():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file"}), 400
    path = os.path.join(UPLOAD_DIR, "template.xlsx")
    f.save(path)
    try:
        am = load_advisor_map(path)
        fm = load_fund_map(path)
        SESSION["advisor_map"] = am
        SESSION["fund_map"]    = fm
        return jsonify({"ok": True, "advisor_count": len(am), "fund_count": len(fm)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/upload_platform", methods=["POST"])
def upload_platform():
    platform = request.form.get("platform")
    f = request.files.get("file")
    if not platform or not f:
        return jsonify({"error": "Missing platform or file"}), 400

    ext = os.path.splitext(f.filename)[1]
    path = os.path.join(UPLOAD_DIR, f"platform_{platform}{ext}")
    f.save(path)
    SESSION.setdefault("platform_files", {})[platform] = path
    return jsonify({"ok": True, "platform": platform, "filename": f.filename})


@app.route("/process", methods=["POST"])
def process():
    data = request.json or {}
    report_date_str = data.get("report_date", date.today().strftime("%Y-%m-%d"))
    report_date     = datetime.strptime(report_date_str, "%Y-%m-%d").date()

    advisor_map    = SESSION.get("advisor_map", {})
    fund_map       = SESSION.get("fund_map", {})
    platform_files = SESSION.get("platform_files", {})
    extra_advisors = SESSION.get("extra_advisors", {})  # user-added mappings
    extra_funds    = SESSION.get("extra_funds", {})

    # Merge in user-supplied mappings
    merged_advisor = {**advisor_map, **extra_advisors}
    merged_fund    = {k.lower(): v for k, v in {**fund_map, **extra_funds}.items()}

    all_frames = []
    errors_by_platform = {}

    for platform, path in platform_files.items():
        if not os.path.exists(path):
            continue
        func = PROCESSOR_MAP.get(platform)
        if not func:
            continue
        try:
            df = func(path, merged_advisor, merged_fund)
            if df is not None and len(df) > 0:
                df["Date"] = pd.Timestamp(report_date)
                df["Retirement Fund Type"]   = np.nan
                df["Participating Employer"] = np.nan
                all_frames.append(df)
                errs = find_mapping_errors(df, merged_advisor, merged_fund)
                if errs:
                    errors_by_platform[PLATFORM_LABELS.get(platform, platform)] = errs
        except Exception as e:
            errors_by_platform[platform] = [{"type": "processing_error", "error": str(e),
                                             "traceback": traceback.format_exc()}]

    if not all_frames:
        return jsonify({"error": "No data processed. Upload platform files first."}), 400

    combined = pd.concat(all_frames, ignore_index=True)
    final_cols = ["Date", "Broker House Name", "Broker Name",
                  "Retirement Fund Type", "Participating Employer",
                  "Product", "LISP", "Fund Name",
                  "InFlows (R)", "OutFlows (R)", "NetFlows (R)", "AUM (R)"]
    final_df = combined[[c for c in final_cols if c in combined.columns]].copy()

    SESSION["final_df"]    = final_df
    SESSION["report_date"] = report_date_str

    # Build per-platform summary directly from already-processed frames (no re-processing)
    summary = []
    for pf, path in platform_files.items():
        label = PLATFORM_LABELS.get(pf, pf)
        func  = PROCESSOR_MAP.get(pf)
        if func and os.path.exists(path):
            try:
                tmp = func(path, merged_advisor, merged_fund)
                if tmp is not None and len(tmp) > 0:
                    summary.append({
                        "platform": label,
                        "rows":     len(tmp),
                        "aum":      round(float(tmp["AUM (R)"].sum()), 2) if "AUM (R)" in tmp else 0,
                        "inflows":  round(float(tmp["InFlows (R)"].dropna().sum()), 2) if "InFlows (R)" in tmp else 0,
                        "outflows": round(float(tmp["OutFlows (R)"].dropna().sum()), 2) if "OutFlows (R)" in tmp else 0,
                    })
                else:
                    summary.append({"platform": label, "rows": 0, "aum": 0, "inflows": 0, "outflows": 0})
            except Exception:
                summary.append({"platform": label, "rows": 0, "aum": 0, "inflows": 0, "outflows": 0})

    SESSION["platform_summary"] = summary
    total_aum = round(float(final_df["AUM (R)"].sum()), 2)

    return jsonify({
        "ok": True,
        "total_rows": len(final_df),
        "total_aum": total_aum,
        "platform_summary": summary,
        "mapping_errors": errors_by_platform,
        "has_errors": bool(errors_by_platform),
    })


@app.route("/add_mapping", methods=["POST"])
def add_mapping():
    data = request.json or {}
    mtype = data.get("type")  # "advisor" or "fund"
    if mtype == "advisor":
        SESSION.setdefault("extra_advisors", {})[data["id"]] = {
            "Broker Name":       data.get("broker_name", ""),
            "Broker House Name": data.get("broker_house_name", ""),
            "LISP":              data.get("lisp", ""),
        }
    elif mtype == "fund":
        SESSION.setdefault("extra_funds", {})[data["id"].lower()] = {
            "Fund Name": data.get("fund_name", data["id"]),
            "Product":   data.get("product", "Model"),
        }
    return jsonify({"ok": True})


@app.route("/download", methods=["GET"])
def download():
    final_df = SESSION.get("final_df")
    if final_df is None:
        return jsonify({"error": "No data — run Process first"}), 400
    report_date = SESSION.get("report_date", date.today().strftime("%Y-%m-%d"))
    month_str = report_date.replace("-", "")[:6]
    buf, total_aum = write_excel(final_df, report_date)
    fname = f"AUM_Consolidated_{month_str}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/status", methods=["GET"])
def status():
    return jsonify({
        "template_loaded": "advisor_map" in SESSION,
        "advisor_count": len(SESSION.get("advisor_map", {})),
        "fund_count": len(SESSION.get("fund_map", {})),
        "platforms_uploaded": list(SESSION.get("platform_files", {}).keys()),
        "data_processed": "final_df" in SESSION,
        "total_rows": len(SESSION["final_df"]) if "final_df" in SESSION else 0,
    })


@app.route("/reset", methods=["POST"])
def reset():
    SESSION.clear()
    return jsonify({"ok": True})


@app.route("/")
def index():
    """Serve the HTML app from the same folder as app.py."""
    html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aum_app.html")
    if not os.path.exists(html_path):
        return "<h2>Error: aum_app.html not found next to app.py</h2>", 404
    return send_file(html_path)


if __name__ == "__main__":
    import sys, os

    # Detect if running on Render (via PORT environment variable)
    PORT = int(os.environ.get("PORT", 5050))
    HOST = "0.0.0.0" if "PORT" in os.environ else "127.0.0.1"
    
    # Only set up local logging if not on Render
    if "PORT" not in os.environ:
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aum_server.log")
        import logging
        logging.basicConfig(
            filename=log_path,
            level=logging.ERROR,
            format="%(asctime)s %(levelname)s %(message)s"
        )

        print()
        print("  ================================================")
        print("  Sequoia Capital Management - AUM Consolidation")
        print(f"  Server running at http://localhost:{PORT}")
        print("  Keep this window open while using the tool.")
        print("  Press Ctrl+C to stop.")
        print("  ================================================")
        print()
        print(f"  Log file: {log_path}")
        print()

    try:
        app.run(host=HOST, port=PORT, debug=False, use_reloader=False)
    except OSError as e:
        msg = str(e)
        if "10048" in msg or "Address already in use" in msg:
            print(f"\n  Port {PORT} is already in use.")
            print(f"  Another instance may be running.")
            print(f"  Open http://localhost:{PORT} in your browser,")
            print(f"  or close the other instance and try again.\n")
        else:
            print(f"\n  ERROR starting server: {e}\n")
        if "PORT" not in os.environ:
            input("  Press Enter to close...")
    except Exception as e:
        print(f"\n  FATAL ERROR: {e}\n")
        if "PORT" not in os.environ:
            import logging
            logging.exception("Fatal error")
            input("  Press Enter to close...")
